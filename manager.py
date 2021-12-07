from json import load

from openpyxl import Workbook
from openpyxl.utils  import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment

class DataManager:
    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = 'Данные'

    def write_columns(self, columns, sizes):
        fill = PatternFill(start_color='5A5A5A',
                           end_color='5A5A5A',
                           fill_type="solid")
        year_fill = PatternFill(start_color='FFC000',
                           end_color='FFC000',
                           fill_type="solid")
        font = Font(color='FFFFFF')
        align = Alignment(horizontal='center', vertical='center', wrapText=True)
        self.ws.row_dimensions[1].height = 50
        for index, value in enumerate(columns, start=1):
            self.ws.column_dimensions[get_column_letter(index)].width = sizes[index - 1]
            if index == 5:
                self.write_cell(1, index, value, fill=year_fill, font=font, align=align)
            else:
                self.write_cell(1, index, value, fill=fill, font=font, align=align)

    def load_data(self, filename):
        f = open(filename)
        return load(f)

    def save_data(self, filename):
        self.wb.save(filename)

    def price_format(self, price, length=3):
        price = str(price)
        res = []
        counter = -length
        res.insert(0, price[counter:])
        while price[counter - length:counter] != '':
            res.insert(0, price[counter - length:counter])
            counter -= length
        return ' '.join(res)

    def write_cell(self, row, column, value, fill=0, font=0, align=0):
        current_cell = self.ws.cell(row=row, column=column)
        current_cell.value = value
        if fill:
            current_cell.fill = fill
        if font:
            current_cell.font = font
        if align:
            current_cell.alignment = align

    def write_data(self, data):
        fill1 = PatternFill(start_color='A5A5A5',
                           end_color='A5A5A5',
                           fill_type="solid")
        fill2 = PatternFill(start_color='5A5A5A',
                           end_color='5A5A5A',
                           fill_type="solid")
        font = Font(color='FFFFFF')
        align = Alignment(horizontal='center')
        pwts = {}
        for index, auto in enumerate(data, start=2):
            if 'first_price' in auto:
                pwts.setdefault(auto['first_price'], dict())
                pwts[auto['first_price']].setdefault(auto['model'], auto['final_price'])
            self.write_cell(index, 1, auto['brand'], fill=fill2, font=font, align=align)
            self.write_cell(index, 2, auto['model'], fill=fill1, font=font, align=align)
            self.write_cell(index, 3, auto['trim'])
            self.write_cell(index, 4, auto['pwt'], align=align)
            self.write_cell(index, 5, auto['year'], align=align)
            self.write_cell(index, 6, self.price_format(auto['price']), align=align)
            self.write_cell(index, 7, auto['discount'], align=align)
            self.write_cell(index, 8, self.price_format(auto['final_price']), align=align)
            self.write_cell(index, 9, auto['utilization'], align=align)
            self.write_cell(index, 10, auto['trade_in'], align=align)
            self.write_cell(index, 11, auto['trade_in_special'], align=align)
            self.write_cell(index, 12, auto['trade_in_local'], align=align)
            self.write_cell(index, 13, auto['direct_discount'], align=align)
            self.write_cell(index, 14, auto['comment'])
        amount = len(data)
        index = 3
        for pwt, models in pwts.items():
            for model, price in models.items():
                self.write_cell(amount + index, 2, model, fill=fill1, font=font, align=align)
                self.write_cell(amount + index, 4, 'Первая цена ' + pwt, align=align)
                self.write_cell(amount + index, 8, self.price_format(price), align=align)
                index += 1
            index += 1